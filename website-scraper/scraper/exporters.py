from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def safe_filename(name: str) -> str:
    value = re.sub(r'[^A-Za-z0-9._-]+', '-', name or 'scrape').strip('-._')
    return (value or 'scrape')[:80]


def _columns(records: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    for record in records:
        for key in record:
            if key not in columns:
                columns.append(str(key))
    return columns


def _cell(value: Any) -> str | int | float | bool:
    if value is None or value == '':
        return 'Not Available'
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def export_all(records: list[dict[str, Any]], folder: str, job_id: int, job_name: str) -> dict[str, str]:
    target = Path(folder)
    target.mkdir(parents=True, exist_ok=True)
    base = f'{job_id}-{safe_filename(job_name)}'
    csv_path = target / f'{base}.csv'
    json_path = target / f'{base}.json'
    xlsx_path = target / f'{base}.xlsx'
    columns = _columns(records)

    with csv_path.open('w', newline='', encoding='utf-8-sig') as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        for row in records:
            writer.writerow({key: _cell(row.get(key, 'Not Available')) for key in columns})

    json_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding='utf-8')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Scraped Data'
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in records:
        sheet.append([_cell(row.get(key, 'Not Available')) for key in columns])
    for index, name in enumerate(columns, start=1):
        values = [len(str(name))]
        for cell in sheet[get_column_letter(index)][1:101]:
            values.append(len(str(cell.value or '')))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(values) + 2, 60)
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)

    return {'csv': str(csv_path), 'json': str(json_path), 'excel': str(xlsx_path)}
